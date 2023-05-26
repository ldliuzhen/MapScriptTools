import re
from openpyxl import Workbook

# 导入所需模块

def find_strings_and_functions(file_path):
    # 定义一个函数，用于在文件中查找字符串和对应的函数
    with open(file_path, 'r', encoding='utf8') as file:
        content = file.readlines()
        # 读取文件内容并将其存储在一个列表中

    result_list = []
    # 初始化一个空列表，用于存储结果

    func_pattern = re.compile(r'(bool|void|int|fixed|string) lib[\dA-F]+_gf_(.*?)\s')
    # 定义一个正则表达式模式，用于匹配函数声明
    # 它匹配类型（bool、void、int、fixed、string）后跟着'lib'和一些十六进制字符，
    # 然后是'_gf_'，并捕获函数名

    search_pattern = re.compile(r'lv_playerHandel\s*==\s*"(.+?)"')
    # 定义一个正则表达式模式，用于搜索特定的字符串模式
    # 它匹配'lv_playerHandel == "some_string"'，并捕获双引号中的字符串

    func_body = ""
    func_name = ""
    brace_count = 0
    # 初始化变量，用于存储函数体、函数名和大括号计数

    for line in content:
        # 遍历文件内容中的每一行
        if brace_count == 0:
            match = func_pattern.search(line)
            # 检查该行是否与函数模式匹配
            if match:
                func_name = match.group(2)
                # 从匹配的模式中获取函数名
                brace_count += line.count("{")
                # 计算该行中开括号'{'的数量，并更新大括号计数
                func_body += line
                # 将该行添加到函数体中
        else:
            brace_count += line.count("{")
            # 计算该行中开括号'{'的数量，并更新大括号计数
            brace_count -= line.count("}")
            # 计算该行中闭括号'}'的数量，并从大括号计数中减去
            func_body += line
            # 将该行添加到函数体中
            if brace_count == 0:
                # 如果大括号计数变为零，表示函数体已结束
                matches = set(search_pattern.findall(func_body))
                # 在函数体中查找所有匹配的字符串模式
                # 使用集合来存储匹配项，以去除重复项
                for match in matches:
                    result_list.append((match, func_name))
                    # 将匹配的字符串和对应的函数名添加到结果列表中
                func_body = ""
                # 重置函数体
                func_name = ""
                # 重置函数名
    
    return result_list
    # 返回匹配的字符串和对应的函数列表

# 调用函数，查找给定文件中的字符串和函数
results = find_strings_and_functions('C:/Users/hnldl/Desktop/MapScript.galaxy')

# 创建一个新的 Excel 工作簿和工作表
wb = Workbook()
ws = wb.active
ws.title = "玩家句柄和函数"

# 将结果写入工作表
for idx, val in enumerate(results, start=1):
    ws.cell(row=idx, column=1, value=val[0])
    # 在第一列写入字符串
    ws.cell(row=idx, column=2, value=val[1])
    # 在第二列写入对应的函数名

# 保存工作簿
wb.save("玩家句柄和函数.xlsx")
